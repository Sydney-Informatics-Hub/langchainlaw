import argparse
import logging
from itertools import chain
import json
from pathlib import Path
from openpyxl import load_workbook, Workbook
import re

from langchainlaw.cache import Cache
from langchainlaw.langchainlaw import parse_llm_json

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

URI_RE = re.compile("https://www.caselaw.nsw.gov.au/decision/([0-9a-f]+)")

MAX_RE = re.compile("context length")

DEFENDANT_RE = re.compile("defendant", flags=re.I)


def load_config(cf_file):
    """Load the config JSON"""
    with open(cf_file, "r") as fh:
        cf = json.load(fh)
        return cf


def expand_ra_cols(cf):
    """Looks for columns CLAIMANT and DEFENDANT in the list of input
    columns amd expands them out to n sets of columns from PARTIES_IN_COLS
    for each category, so we end up with

    claimant_1_relationship_to_party
    claimant_1_is_dependant
    [...]
    claimant_2_relationship_to_party
    claimant_2_is_dependant
    [...]

    and so on
    """

    base = [[c] for c in cf["SPREADSHEET_IN_COLS"]]
    party_cols = cf["PARTIES_IN_COLS"]
    n = cf["PARTIES_N"]
    for p in ["CLAIMANT", "DEFENDANT"]:
        partytype = p.lower()
        i = base.index([p])
        base[i] = [f"{partytype}_{i + 1}_{col}" for i in range(n) for col in party_cols]
    return list(chain.from_iterable(base))  # flattens list of lists


def load_ra_spreadsheet(config):
    """Load the spreadsheet with the RA's summary of the cases

    Returns a dict of ID -> list of spreadsheet rows
    """
    cases = {}
    header = True
    cols = expand_ra_cols(config)
    wb = load_workbook(config["SPREADSHEET_IN"])
    for row in wb.active:
        if header:
            header = False
        else:
            case = dict(zip(cols, [cell.value for cell in row]))
            case["id"] = parse_case_uri(case["uri"])
            if case["id"]:
                if case["id"] not in cases:
                    cases[case["id"]] = [case]
                else:
                    if not case["mnc"]:
                        # mnc cells are merged so get it from the first row
                        case["mnc"] = cases[case["id"]][0]["mnc"]
                    cases[case["id"]].append(case)
            else:
                uri = case["uri"]
                row_i = row[0].row
                logger.warning(f"Row [{row_i}]: Couldn't parse case ID from {uri}")
    return cols, cases


def parse_case_uri(uri):
    """Try to get a case ID from a URI in the input spreadsheet"""
    if uri is None:
        return None
    m = URI_RE.match(uri)
    if m:
        return m.group(1)
    return None


def find_cached_results(cache, caseid, mapping):
    """Look up a returned value from the LLM in the cache. Returns None if
    the case isn't in the cache or if the LLM request failed."""
    mapped = {}
    if not cache.exists(caseid):
        logger.debug(f"Case {caseid} not in cache")
        return None
    for field, spreadsheet in mapping.items():
        try:
            mapped[field] = cache.read(caseid, field)
            if MAX_RE.search(mapped[field]):
                logger.warning(f"Case {caseid} exceeded token length")
                return None
        except Exception as e:
            mapped[field] = f"Cache read failed: {e}"

    return mapped


def make_headers(out_cols):
    """Take the output columns and return a list of column headings by
    expanding the keys for multivalue columns"""
    header = []
    subhead = []
    for col, mapping in out_cols.items():
        header.append(col)
        if type(mapping) is str:
            subhead.append("")
        else:
            subheads = list(mapping.keys())
            header += ["" for _ in subheads[1:]]
            subhead += subheads
    return header, subhead


def add_ra_parties(ws, row, col, ra_parties):
    """Hack to expand the 'parties' field in the RA spreadsheet into the
    multiple rows in the collated spreadsheet"""
    json_parties = ra_parties.replace("'", '"')
    try:
        parties = json.loads(json_parties)
        ws.cell(row=row, column=col).value = parties[0]
        if len(parties) > 2:
            logger.warning("case has more parties than expected:")
            logger.warning(str(parties))
            ws.cell(row=row + 1, column=col).value = ",".join(parties[1:])
        else:
            ws.cell(row=row + 1, column=col).value = parties[1]
    except Exception as e:
        logger.warning("Expanding RA parties failed")
        logger.warning(ra_parties)
        logger.warning(str(e))


def guess_party(llm_party):
    if DEFENDANT_RE.search(llm_party):
        return "defendant"
    else:
        return "claimant"


def get_from_party(party, field):
    """Fixes relationship_to_party -> relationship_to_deceased based on
    an error in the prompt"""

    val = party.get(field, "")
    if not val and field == "relationship_to_party":
        val = party.get("relationship_to_deceased")
    return val


def flatten_llm_result(in_cols, mappings, llm_results):
    """Writes most of the llm results straight into the output columns.
    For multivalue ones: parties get distributed across the n claimants and
    n defendants columns, all other multivalues just get written as a JSON
    dump to one column"""
    llm_vals = {}
    party_names = []
    for col, mapping in mappings.items():
        if col == "parties":
            parties_n = {"claimant": 0, "defendant": 0}
            parties = parse_llm_json(llm_results[col])
            for party in parties:
                party_type = guess_party(party["role_in_trial"])
                parties_n[party_type] += 1
                prefix = party_type + "_" + str(parties_n[party_type])
                party_names.append(party["name"])
                for llm_col, ra_col in mapping.items():
                    if ra_col is not None:
                        llm_vals[f"{prefix}_{ra_col}"] = get_from_party(party, llm_col)
        else:
            if type(mapping) is str:
                llm_vals[mapping] = llm_results[col]
            else:
                # assumes that for all multivalues apart from parties, col is
                # in out_cols
                llm_json = parse_llm_json(llm_results[col])  # unquote it
                llm_vals[col] = json.dumps(llm_json)
    llm_vals["parties"] = json.dumps(party_names)
    return [llm_vals.get(c, "-") for c in in_cols]


def dump_cases(ra_cases):
    """Dump a precis of the ra_cases to check that the load is working"""
    for case_id, cases in ra_cases.items():
        for case in cases:
            mnc = case["mnc"]
            uri = case["uri"]
            title = case["title"]
            ra = case["RA"]
            print(f"{case_id},{ra},{uri},{mnc},{title}")


def test_flatten(cf):
    """test flattening on cases which aren't in the RA spreadsheet"""
    cache_dir = Path(cf["CACHE"])
    cache = Cache(cf["CACHE"])
    mapping = cf["SPREADSHEET_OUT_COLS"]
    full_cols = expand_ra_cols(cf)
    for item in cache_dir.glob("*"):
        if item.is_dir():
            case_id = item.name
            print(case_id)
            llm_results = find_cached_results(cache, case_id, mapping)
            if llm_results is not None:
                llm_cols = flatten_llm_result(full_cols, mapping, llm_results)
                print(llm_cols)


def collate():
    ap = argparse.ArgumentParser("collate-langchain")
    ap.add_argument(
        "--config",
        default="./collate_config.json",
        type=Path,
        help="Config file",
    )
    ap.add_argument(
        "--flatten",
        action="store_true",
        default=False,
        help="Flatten multiple results into a single row",
    )
    args = ap.parse_args()
    cf = load_config(args.config)
    cols, ra_cases = load_ra_spreadsheet(cf)
    mappings = cf["SPREADSHEET_OUT_COLS"]
    cache = Cache(cf["CACHE"])
    results = Workbook()
    ws = results.active
    ws.append(cols)
    for case_id, ra_case in ra_cases.items():
        logger.warning(case_id)
        for ra_row in ra_case:
            ws.append([ra_row[c] for c in cols])
        llm_results = find_cached_results(cache, case_id, mappings)
        if llm_results is not None:
            llm_cols = flatten_llm_result(cols, mappings, llm_results)
            for i in range(11):
                llm_cols[i] = ra_case[0][cols[i]]
            llm_cols[1] = "GPT-4o"
            ws.append(llm_cols)
        else:
            ws.append([ra_case[0][1], "GPT-4o", "No results"])
    results.save(cf["SPREADSHEET_OUT"])
    logger.warning("Wrote collated results to " + cf["SPREADSHEET_OUT"])


if __name__ == "__main__":
    collate()
